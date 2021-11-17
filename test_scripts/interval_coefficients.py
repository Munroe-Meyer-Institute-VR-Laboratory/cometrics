import json
import os
import pathlib
from os import walk, path
import openpyxl
import csv
import datetime
from tkinter import *
from tkinter import filedialog, messagebox


def cal_acc(rel_filename, prim_filename, ksf_filename, window_size):
    if path.isfile(rel_filename) and path.isfile(prim_filename) and path.isfile(ksf_filename):
        try:
            with open(ksf_filename) as f:
                keystroke_json = json.load(f)
            freq_bindings = []
            dur_bindings = []
            for key in keystroke_json:
                if key == "Frequency":
                    for bindings in keystroke_json[key]:
                        freq_bindings.append(bindings[0])
                if key == "Duration":
                    for bindings in keystroke_json[key]:
                        dur_bindings.append(bindings[0])
            with open(prim_filename, 'r') as f:
                prim_session = json.load(f)
            with open(rel_filename, 'r') as f:
                rel_session = json.load(f)
            prim_window_freq, prim_window_dur = get_keystroke_window(ksf_filename, prim_session, window_size)
            rel_window_freq, rel_window_dur = get_keystroke_window(ksf_filename, rel_session, window_size)

            if len(prim_window_freq) > len(rel_window_freq):
                for i in range(0, len(prim_window_freq) - len(rel_window_freq)):
                    rel_window_freq.append([0] * len(freq_bindings))
            elif len(prim_window_freq) < len(rel_window_freq):
                for i in range(0, len(rel_window_freq) - len(prim_window_freq)):
                    prim_window_freq.append([0] * len(freq_bindings))
            if len(prim_window_dur) > len(rel_window_dur):
                for i in range(0, len(prim_window_dur) - len(rel_window_dur)):
                    rel_window_dur.append([0] * len(dur_bindings))
            elif len(prim_window_dur) < len(rel_window_dur):
                for i in range(0, len(rel_window_dur) - len(prim_window_dur)):
                    prim_window_dur.append([0] * len(dur_bindings))
            freq_pia = [0] * len(freq_bindings)
            freq_oia_agree, freq_oia_disagree = [0] * len(freq_bindings), [0] * len(freq_bindings)
            freq_nia_agree, freq_nia_disagree = [0] * len(freq_bindings), [0] * len(freq_bindings)
            freq_eia_agree = [0] * len(freq_bindings)
            freq_tia_agree = [0] * len(freq_bindings)
            freq_intervals, dur_intervals = [0] * len(freq_bindings), [0] * len(dur_bindings)
            dur_pia = [0] * len(dur_bindings)
            dur_eia_agree = [0] * len(dur_bindings)
            for row in range(0, len(prim_window_freq)):
                for cell in range(0, len(prim_window_freq[row])):
                    if prim_window_freq[row][cell] > rel_window_freq[row][cell]:
                        larger = prim_window_freq[row][cell]
                        smaller = rel_window_freq[row][cell]
                    else:
                        smaller = prim_window_freq[row][cell]
                        larger = rel_window_freq[row][cell]
                    if smaller == larger:
                        x = 1
                    else:
                        x = smaller / larger
                    freq_pia[cell] += x
                    if larger == 0 and smaller == 0 or larger >= 1 and smaller >= 1:
                        freq_tia_agree[cell] += 1
                    if larger == smaller:
                        freq_eia_agree[cell] += 1

                    if larger == 0 and smaller == 0:
                        freq_nia_agree[cell] += 1
                    elif larger >= 1 and smaller == 0:
                        freq_nia_disagree[cell] += 1

                    if larger >= 1 and smaller >= 1:
                        freq_oia_agree[cell] += 1
                    elif larger >= 1 and smaller == 0:
                        freq_oia_disagree[cell] += 1

                    freq_intervals[cell] += 1
            for row in range(0, len(prim_window_dur)):
                for cell in range(0, len(prim_window_dur[row])):
                    if prim_window_dur[row][cell] > rel_window_dur[row][cell]:
                        larger = prim_window_dur[row][cell]
                        smaller = rel_window_dur[row][cell]
                    else:
                        smaller = prim_window_dur[row][cell]
                        larger = rel_window_dur[row][cell]
                    if smaller == larger:
                        x = 1
                        dur_eia_agree[cell] += 1
                    else:
                        x = smaller / larger
                    dur_pia[cell] += x
                    dur_intervals[cell] += 1
            headers = {
                "Generation Date": datetime.datetime.today().strftime
                    ("%B %d, %Y") + " " + datetime.datetime.now().strftime("%H:%M:%S"),
                "Primary Session Date": prim_session["Session Date"] + " " + prim_session["Session Start Time"],
                "Primary Therapist": prim_session["Primary Therapist"],
                "Primary Case Manager": prim_session["Case Manager"],
                "Primary Session Therapist": prim_session["Session Therapist"],
                "Reliability Session Date": rel_session["Session Date"] + " " + rel_session["Session Start Time"],
                "Reliability Therapist": rel_session["Primary Therapist"],
                "Reliability Case Manager": rel_session["Case Manager"],
                "Reliability Session Therapist": rel_session["Session Therapist"],
                "Window Size (seconds)": str(window_size) + " seconds"
            }
            path_to_file = filedialog.asksaveasfilename(
                defaultextension='.xlsx', filetypes=[("Excel files", '*.xlsx')],
                title="Choose output filename")
            wb = openpyxl.Workbook()
            ws = wb.active
            wb.create_sheet("Primary Data")
            wb.create_sheet("Reliability Data")
            ws.title = "Interval Measures"
            for x, header in zip(range(1, 10), headers):
                ws.cell(row=x, column=1).value = header
                ws.cell(row=x, column=2).value = headers[header]

            row = 11
            for col, val in enumerate(freq_bindings, start=2):
                ws.cell(row=row, column=col).value = val
            row += 1
            ws.cell(row=row, column=1).value = "Freq PIA"
            for col, val in enumerate(freq_pia, start=2):
                ws.cell(row=row, column=col).value = str(int((val / freq_intervals[freq_pia.index(val)]) * 100)) + "%"
            row += 1
            ws.cell(row=row, column=1).value = "Freq NIA"
            for col, val in enumerate(freq_nia_agree, start=2):
                ws.cell(row=row, column=col).value = str \
                    (int((val / (val + freq_nia_disagree[freq_nia_agree.index(val)])) * 100)) + "%"
            row += 1
            ws.cell(row=row, column=1).value = "Freq TIA"
            for col, val in enumerate(freq_tia_agree, start=2):
                ws.cell(row=row, column=col).value = str \
                    (int((val / freq_intervals[freq_tia_agree.index(val)]) * 100)) + "%"
            row += 1
            ws.cell(row=row, column=1).value = "Freq EIA"
            for col, val in enumerate(freq_eia_agree, start=2):
                ws.cell(row=row, column=col).value = str \
                    (int((val / freq_intervals[freq_eia_agree.index(val)]) * 100)) + "%"
            row += 1
            ws.cell(row=row, column=1).value = "Freq OIA"
            for col, val in enumerate(freq_oia_agree, start=2):
                ws.cell(row=row, column=col).value = str \
                    (int((val / (val + freq_oia_disagree[freq_oia_agree.index(val)])) * 100)) + "%"
            row += 2
            for col, val in enumerate(dur_bindings, start=2):
                ws.cell(row=row, column=col).value = val
            row += 1
            ws.cell(row=row, column=1).value = "Dur PIA"
            for col, val in enumerate(dur_pia, start=2):
                ws.cell(row=row, column=col).value = str(int((val / dur_intervals[dur_pia.index(val)]) * 100)) + "%"
            row += 1
            ws.cell(row=row, column=1).value = "Dur EIA"
            for col, val in enumerate(dur_eia_agree, start=2):
                ws.cell(row=row, column=col).value = str \
                    (int((val / dur_intervals[dur_eia_agree.index(val)]) * 100)) + "%"
            row += 1

            ws = wb["Primary Data"]
            row = 1
            for col, val in enumerate(freq_bindings, start=1):
                ws.cell(row=row, column=col).value = val
            row += 1
            for window in prim_window_freq:
                for col, val in enumerate(window, start=1):
                    ws.cell(row=row, column=col).value = val
                row += 1
            row += 1
            for col, val in enumerate(dur_bindings, start=1):
                ws.cell(row=row, column=col).value = val
            row += 1
            for window in prim_window_dur:
                for col, val in enumerate(window, start=1):
                    ws.cell(row=row, column=col).value = val
                row += 1

            ws = wb["Reliability Data"]
            row = 1
            for col, val in enumerate(freq_bindings, start=1):
                ws.cell(row=row, column=col).value = val
            row += 1
            for window in rel_window_freq:
                for col, val in enumerate(window, start=1):
                    ws.cell(row=row, column=col).value = val
                row += 1
            row += 1
            for col, val in enumerate(dur_bindings, start=1):
                ws.cell(row=row, column=col).value = val
            row += 1
            for window in rel_window_dur:
                for col, val in enumerate(window, start=1):
                    ws.cell(row=row, column=col).value = val
                row += 1

            wb.save(path_to_file)
            os.startfile(pathlib.Path(path_to_file).parent)
            # self.root.iconify()
        except Exception as e:
            messagebox.showerror("Error", "Error encountered!\n" + str(e))
    else:
        messagebox.showwarning("Warning", "Please choose valid files!")


def get_freq_pia(prim_windows, rel_windows):
    freq_pia = [[]]
    for row in range(0, len(prim_windows)):
        for cell in range(0, len(prim_windows[row])):
            if prim_windows[row][cell] > rel_windows[row][cell]:
                larger = prim_windows[row][cell]
                smaller = rel_windows[row][cell]
            else:
                smaller = prim_windows[row][cell]
                larger = rel_windows[row][cell]
            if smaller == 0 and larger == 0:
                x = 1
            else:
                x = smaller / larger


def get_freq_oia(prim_windows, rel_windows):
    for row in range(0, len(prim_windows)):
        for cell in range(0, len(prim_windows[row])):
            pass


def get_freq_nia(prim_windows, rel_windows):
    for row in range(0, len(prim_windows)):
        for cell in range(0, len(prim_windows[row])):
            pass


def get_freq_eia(prim_windows, rel_windows):
    for row in range(0, len(prim_windows)):
        for cell in range(0, len(prim_windows[row])):
            pass


def get_freq_tia(prim_windows, rel_windows):
    for row in range(0, len(prim_windows)):
        for cell in range(0, len(prim_windows[row])):
            pass


def get_dur_pia(prim_windows, rel_windows):
    for row in range(0, len(prim_windows)):
        for cell in range(0, len(prim_windows[row])):
            pass


def get_dur_eia(prim_windows, rel_windows):
    for row in range(0, len(prim_windows)):
        for cell in range(0, len(prim_windows[row])):
            pass


def get_keystroke_window(key_file, session_file, window_size):
    session_time = int(session_file["Session Time"])
    with open(key_file) as f:
        keystroke_json = json.load(f)
    freq_bindings = []
    dur_bindings = []
    for key in keystroke_json:
        if key == "Frequency":
            for bindings in keystroke_json[key]:
                freq_bindings.append(bindings[0])
        if key == "Duration":
            for bindings in keystroke_json[key]:
                dur_bindings.append(bindings[0])
    if session_time % window_size != 0:
        session_time += window_size - (session_time % window_size)
    freq_windows = [[0] * len(freq_bindings) for i in range(int(session_time / window_size))]
    dur_windows = [[0] * len(dur_bindings) for i in range(int(session_time / window_size))]
    keys = list(session_file.keys())[12:]
    dur_keys, freq_keys = [], []
    for key in keys:
        if type(session_file[key][1]) is list:
            dur_keys.append(key)
        else:
            freq_keys.append(key)
    for i in range(0, session_time, window_size):
        for key in freq_keys[:]:
            if i <= session_file[key][1] < (i + window_size):
                freq_windows[int(i / window_size)][freq_bindings.index(session_file[key][0])] += 1
                freq_keys.remove(key)
            else:
                break
    for i in range(0, session_time, window_size):
        for key in dur_keys[:]:
            if i <= session_file[key][1][0] < (i + window_size):
                dur_windows[int(i / window_size)][dur_bindings.index(session_file[key][0])] += 1
                if i <= session_file[key][1][1] < (i + window_size):
                    dur_keys.remove(key)
            elif i <= session_file[key][1][1] < (i + window_size):
                dur_windows[int(i / window_size)][dur_bindings.index(session_file[key][0])] += 1
                dur_keys.remove(key)
            else:
                break
    return freq_windows, dur_windows


rel = r'D:\GitHub Repos\KeystrokeAnnotator\experiments\Lego Master\data\Tucan Sam\sessions\session_1.json'
prim = r'D:\GitHub Repos\KeystrokeAnnotator\experiments\Lego Master\data\Tucan Sam\sessions\session_2.json'
ksf_file = r'D:\GitHub Repos\KeystrokeAnnotator\experiments\Lego Master\data\Tucan Sam\keystrokes\Reference_Tracker.json'
window_size = 10

cal_acc(rel, rel, ksf_file, window_size)
