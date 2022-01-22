import csv
import datetime
import glob
import json
import os
import pathlib
import traceback
from os import walk
from shutil import copy2
from tkinter import messagebox

import openpyxl
from openpyxl import Workbook
from openpyxl.cell import MergedCell, Cell

from logger_util import *


def import_ksf(filename, ksf_dir):
    if filename:
        try:
            wb = openpyxl.load_workbook(filename)
            data_wb = None
            try:
                data_wb = wb['Data']
            except KeyError as e:
                messagebox.showerror("Error", "Sheet 'Data' not found!\n"
                                              "This is required for importing the keystrokes!\n" + str(e))
                print("ERROR: Sheet 'Data' not found!\n"
                      "This is required for importing the keystrokes!\n" + str(e))
            if data_wb is not None:
                tracker_headers, freq_headers, dur_headers = get_key_cells(data_wb)
                conditions, freq_keys, dur_keys = [], [], []
                for key in freq_headers:
                    freq_keys.append([freq_headers[key][2], key])
                for key in dur_headers:
                    dur_keys.append([dur_headers[key][2], key])
                try:
                    cond_wb = wb['Conditions']
                    for i in range(1, 20):
                        if cond_wb['A' + str(i)].value is None:
                            break
                        else:
                            conditions.append(str(cond_wb['A' + str(i)].value))
                except KeyError as e:
                    print("No conditions in the workbook!", str(e))
                    messagebox.showwarning("Warning", "Conditions not found in selected spreadsheet!\n"
                                                      "Add sheet called 'Conditions' for this feature!")
                    print("ERROR: Conditions not found in selected spreadsheet!\n"
                          "Add sheet called 'Conditions' for this feature!")
                name = pathlib.Path(filename).stem
                x = {
                    "Name": name,
                    "Frequency": freq_keys,
                    "Duration": dur_keys,
                    "Conditions": conditions
                }
                keystroke_file = path.join(ksf_dir, name + '.json')
                original_move = path.join(ksf_dir, pathlib.Path(filename).name)
                with open(path.normpath(keystroke_file), 'w') as f:
                    json.dump(x, f)
                if filename != original_move:
                    if path.exists(original_move):
                        os.remove(original_move)
                    copy2(filename, original_move)
                return keystroke_file, x
        except Exception as e:
            messagebox.showwarning("Warning", "Excel format is not correct!\n" + str(e))
            print(f"WARNING: Excel format is not correct\n{str(e)}\n{traceback.print_exc()}")
    else:
        messagebox.showwarning("Warning", "Select the protocol Excel tracker spreadsheet.")
        print("WARNING: Select the protocol Excel tracker spreadsheet")


def load_ksf(ksf_filename):
    with open(ksf_filename) as f:
        keystroke_json = json.load(f)
    freq_bindings = []
    dur_bindings = []
    for key in keystroke_json:
        if key == "Frequency":
            for bindings in keystroke_json[key]:
                freq_bindings.append(bindings[0])
        if key == "Duration":
            for bindings in keystroke_json[key]:
                dur_bindings.append(bindings[0])
    return freq_bindings, dur_bindings


def cal_acc(prim_filename, reli_filename, window_size, output_dir):
    if path.isfile(reli_filename) and path.isfile(prim_filename):
        try:
            # Load in prim and reli files
            with open(prim_filename, 'r') as f:
                prim_session = json.load(f)
            with open(reli_filename, 'r') as f:
                reli_session = json.load(f)
            prim_ksf, reli_ksf, reli_type = prim_session["Keystroke File"], reli_session["Keystroke File"], \
                                            reli_session["Primary Data"]
            freq_b = prim_session["KSF"]["Frequency"]
            freq_bindings = []
            for freq in freq_b:
                freq_bindings.append(freq[1])
            dur_b = prim_session["KSF"]["Duration"]
            dur_bindings = []
            for dur in dur_b:
                dur_bindings.append(dur[1])
            prim_num, reli_num = int(prim_session["Session Number"]), int(reli_session["Session Number"])
            # Perform error checking before causing errors
            if prim_num != reli_num:
                messagebox.showerror("Error", "Session numbers are not the same!")
                print("ERROR: Session numbers are not the same")
                return
            if prim_ksf != reli_ksf:
                messagebox.showerror("Error", "Sessions do not use the same KSF file!")
                print("ERROR: Sessions do not use the same KSF file")
                return
            elif reli_type == "Primary":
                messagebox.showerror("Error", "Selected reliability file is not a reliability collection!")
                print("ERROR: Selected reliability file is not a reliability collection")
                return
            # Generate freq and dur windows
            prim_window_freq, prim_window_dur = get_keystroke_window(freq_bindings, dur_bindings,
                                                                     prim_session, int(window_size))
            rel_window_freq, rel_window_dur = get_keystroke_window(freq_bindings, dur_bindings,
                                                                   reli_session, int(window_size))
            # Append windows to accommodate jagged sessions
            if len(prim_window_freq) > len(rel_window_freq):
                for i in range(0, len(prim_window_freq) - len(rel_window_freq)):
                    rel_window_freq.append([0] * len(freq_bindings))
            elif len(prim_window_freq) < len(rel_window_freq):
                for i in range(0, len(rel_window_freq) - len(prim_window_freq)):
                    prim_window_freq.append([0] * len(freq_bindings))
            if len(prim_window_dur) > len(rel_window_dur):
                for i in range(0, len(prim_window_dur) - len(rel_window_dur)):
                    rel_window_dur.append([0] * len(dur_bindings))
            elif len(prim_window_dur) < len(rel_window_dur):
                for i in range(0, len(rel_window_dur) - len(prim_window_dur)):
                    prim_window_dur.append([0] * len(dur_bindings))
            # Setup IOA variables
            freq_pia = [0] * len(freq_bindings)
            freq_oia_agree, freq_oia_disagree = [0] * len(freq_bindings), [0] * len(freq_bindings)
            freq_nia_agree, freq_nia_disagree = [0] * len(freq_bindings), [0] * len(freq_bindings)
            freq_eia_agree = [0] * len(freq_bindings)
            freq_tia_agree = [0] * len(freq_bindings)
            freq_intervals, dur_intervals = [0] * len(freq_bindings), [0] * len(dur_bindings)
            dur_pia = [0] * len(dur_bindings)
            dur_eia_agree = [0] * len(dur_bindings)
            # Iterate through the windows
            for row in range(0, len(prim_window_freq)):
                for cell in range(0, len(prim_window_freq[row])):
                    if prim_window_freq[row][cell] > rel_window_freq[row][cell]:
                        larger = prim_window_freq[row][cell]
                        smaller = rel_window_freq[row][cell]
                    else:
                        smaller = prim_window_freq[row][cell]
                        larger = rel_window_freq[row][cell]
                    if smaller == larger:
                        x = 1
                    else:
                        x = smaller / larger
                    freq_pia[cell] += x
                    if larger == 0 and smaller == 0 or larger >= 1 and smaller >= 1:
                        freq_tia_agree[cell] += 1
                    if larger == smaller:
                        freq_eia_agree[cell] += 1

                    if larger == 0 and smaller == 0:
                        freq_nia_agree[cell] += 1
                    elif larger >= 1 and smaller == 0:
                        freq_nia_disagree[cell] += 1

                    if larger >= 1 and smaller >= 1:
                        freq_oia_agree[cell] += 1
                    elif larger >= 1 and smaller == 0:
                        freq_oia_disagree[cell] += 1

                    freq_intervals[cell] += 1
            for row in range(0, len(prim_window_dur)):
                for cell in range(0, len(prim_window_dur[row])):
                    if prim_window_dur[row][cell] > rel_window_dur[row][cell]:
                        larger = prim_window_dur[row][cell]
                        smaller = rel_window_dur[row][cell]
                    else:
                        smaller = prim_window_dur[row][cell]
                        larger = rel_window_dur[row][cell]
                    if smaller == larger:
                        x = 1
                        dur_eia_agree[cell] += 1
                    else:
                        x = smaller / larger
                    dur_pia[cell] += x
                    dur_intervals[cell] += 1
            headers = {
                "Generation Date": datetime.datetime.today().strftime
                                   ("%B %d, %Y") + " " + datetime.datetime.now().strftime("%H:%M:%S"),
                "Primary Session Date": prim_session["Session Date"] + " " + prim_session["Session Start Time"],
                "Primary Therapist": prim_session["Primary Therapist"],
                "Primary Case Manager": prim_session["Case Manager"],
                "Primary Session Therapist": prim_session["Session Therapist"],
                "Reliability Session Date": reli_session["Session Date"] + " " + reli_session["Session Start Time"],
                "Reliability Therapist": reli_session["Primary Therapist"],
                "Reliability Case Manager": reli_session["Case Manager"],
                "Reliability Session Therapist": reli_session["Session Therapist"],
                "Window Size (seconds)": str(window_size) + " seconds"
            }
            path_to_file = path.join(output_dir, f"{prim_ksf}_IOA_{prim_num}.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            wb.create_sheet("Primary Data")
            wb.create_sheet("Reliability Data")
            ws.title = "Interval Measures"
            for x, header in zip(range(1, 10), headers):
                ws.cell(row=x, column=1).value = header
                ws.cell(row=x, column=2).value = headers[header]

            row = 11
            for col, val in enumerate(freq_bindings, start=2):
                ws.cell(row=row, column=col).value = val
            row += 1
            ws.cell(row=row, column=1).value = "Freq PIA"
            for col, val in enumerate(freq_pia, start=2):
                ws.cell(row=row, column=col).value = str(int((val / freq_intervals[col - 2]) * 100)) + "%"
            row += 1
            ws.cell(row=row, column=1).value = "Freq NIA"
            for col, val in enumerate(freq_nia_agree, start=2):
                # This protects from when there are no occurrences at all in the session
                if val == 0 and freq_nia_disagree[col - 2] == 0:
                    ws.cell(row=row, column=col).value = "N/A"
                else:
                    ws.cell(row=row, column=col).value = str(int((val / (val + freq_nia_disagree[col - 2])) * 100)) + "%"
            row += 1
            ws.cell(row=row, column=1).value = "Freq TIA"
            for col, val in enumerate(freq_tia_agree, start=2):
                ws.cell(row=row, column=col).value = str(int((val / freq_intervals[col - 2]) * 100)) + "%"
            row += 1
            ws.cell(row=row, column=1).value = "Freq EIA"
            for col, val in enumerate(freq_eia_agree, start=2):
                ws.cell(row=row, column=col).value = str(int((val / freq_intervals[col - 2]) * 100)) + "%"
            row += 1
            ws.cell(row=row, column=1).value = "Freq OIA"
            for col, val in enumerate(freq_oia_agree, start=2):
                # This protects from when there are no occurrences at all in the session
                if val == 0 and freq_oia_disagree[col - 2] == 0:
                    ws.cell(row=row, column=col).value = "N/A"
                else:
                    ws.cell(row=row, column=col).value = str(int((val / (val + freq_oia_disagree[col - 2])) * 100)) + "%"
            row += 2
            for col, val in enumerate(dur_bindings, start=2):
                ws.cell(row=row, column=col).value = val
            row += 1
            ws.cell(row=row, column=1).value = "Dur PIA"
            for col, val in enumerate(dur_pia, start=2):
                ws.cell(row=row, column=col).value = str(int((val / dur_intervals[col - 2]) * 100)) + "%"
            row += 1
            ws.cell(row=row, column=1).value = "Dur EIA"
            for col, val in enumerate(dur_eia_agree, start=2):
                ws.cell(row=row, column=col).value = str(int((val / dur_intervals[col - 2]) * 100)) + "%"
            row += 1

            ws = wb["Primary Data"]
            row = 1
            for col, val in enumerate(freq_bindings, start=1):
                ws.cell(row=row, column=col).value = val
            row += 1
            for window in prim_window_freq:
                for col, val in enumerate(window, start=1):
                    ws.cell(row=row, column=col).value = val
                row += 1
            row += 1
            for col, val in enumerate(dur_bindings, start=1):
                ws.cell(row=row, column=col).value = val
            row += 1
            for window in prim_window_dur:
                for col, val in enumerate(window, start=1):
                    ws.cell(row=row, column=col).value = val
                row += 1

            ws = wb["Reliability Data"]
            row = 1
            for col, val in enumerate(freq_bindings, start=1):
                ws.cell(row=row, column=col).value = val
            row += 1
            for window in rel_window_freq:
                for col, val in enumerate(window, start=1):
                    ws.cell(row=row, column=col).value = val
                row += 1
            row += 1
            for col, val in enumerate(dur_bindings, start=1):
                ws.cell(row=row, column=col).value = val
            row += 1
            for window in rel_window_dur:
                for col, val in enumerate(window, start=1):
                    ws.cell(row=row, column=col).value = val
                row += 1

            wb.save(path_to_file)
            return path_to_file
        except Exception as e:
            messagebox.showerror("Error", "Error encountered!\n" + str(e))
            print(f"ERROR: Error encountered:\n{str(e)}\n{traceback.print_exc()}")
    else:
        messagebox.showwarning("Warning", "Please choose valid files!")


def export_columnwise_csv(prim_dir, reli_dir, output_dir):
    # Get existing session files
    prim_sessions = get_session_files(prim_dir)
    reli_sessions = get_session_files(reli_dir)
    # Create the primary and reliability directories
    prim_export = path.join(output_dir, "Primary")
    reli_export = path.join(output_dir, "Reliability")
    prim_export_files = []
    reli_export_files = []
    # Create directories if they don't exist
    if not path.exists(output_dir):
        os.mkdir(output_dir)
    if not path.exists(prim_export):
        os.mkdir(prim_export)
    else:
        prim_export_files = get_export_files(prim_export)
    # Get reliability files from export dir, if it exists
    if not path.exists(reli_export):
        os.mkdir(reli_export)
    else:
        reli_export_files = get_export_files(reli_export)
    # Convert files
    convert_json_csv(prim_sessions, prim_export_files, prim_export)
    convert_json_csv(reli_sessions, reli_export_files, reli_export)


def convert_json_csv(json_files, existing_files, output_dir):
    # Get the current time to put into export files
    date = datetime.datetime.today().strftime("%B %d, %Y")
    time = datetime.datetime.now().strftime("%H:%M:%S")
    # Iterate through files
    for file in json_files:
        # If converted file exists already, skip it
        name = pathlib.Path(file).stem
        if name in existing_files:
            continue
        # Load session and split it up
        with open(file, 'r') as f:
            session = json.load(f)
        session_data = {k: v for k, v in session.items() if k in list(session.keys())[:15]}
        event_history = session["Event History"]
        # TODO: Export E4 data to CSV... somehow
        e4_data = session["E4 Data"]
        # Open output file and write session to it
        with open(path.join(output_dir, f"{name}.csv"), 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow([name, 'Generated on', date, time])
            # Write out the session fields
            writer.writerow(['Field', 'Value'])
            for key in session_data:
                writer.writerow([key, session[key]])
            # Write out the event history
            writer.writerow(['Tag', 'Onset', 'Offset', 'Frame', 'E4 Window'])
            for event in event_history:
                row = [event[0]]
                if type(event[1]) is list:
                    row.append(event[1][0])
                    row.append(event[1][1])
                else:
                    row.append(event[1])
                    row.append('')
                row.append(event[2])
                row.append(event[3])
                writer.writerow(row)


def get_key_cells(data_wb):
    tracker_headers = {'Assessment:____________': [Cell, '', '', 3],
                       'Client:________': [Cell, '', '', 2],
                       'Data Coll.': [Cell, '', '', 2],
                       'Therapist': [Cell, '', '', 1],
                       'Session': [Cell, '', '', 1],
                       'Cond.': [Cell, '', '', 1],
                       'Date': [Cell, '', '', 1],
                       'Primary': [Cell, '', '', 1],
                       'Reliability': [Cell, '', '', 1],
                       'Notes': [Cell, '', '', 1],
                       'Sess. Dur. (mins)': [Cell, '', '', 1],
                       'Session Data': [Cell, '', '', 1],
                       'Session Data Start': None,
                       'Frequency': [Cell, '', '', 1],
                       'Frequency Start': None,
                       'Duration': [Cell, '', '', 1],
                       'Duration Start': None,
                       'ST': [Cell, '', '', 1],
                       'PT': [Cell, '', '', 1],
                       'Session Time': [Cell, '', '', 1],
                       'Session Time Start': None,
                       'Pause Time': [Cell, '', '', 1],
                       'Pause Time Start': None
                       }
    freq_headers = {}
    dur_headers = {}
    # Parse the tracker file and find the key cells
    for row in data_wb.iter_rows(1, 4):
        for i in range(0, len(row)):
            if tracker_headers['Frequency Start'] and tracker_headers['Duration Start']:
                if row[i].value:
                    if len(row[i].value) == 1:
                        col_value = ''.join([i for i in row[i].coordinate if not i.isdigit()])
                        value_cell = ''.join([i for i in row[i].coordinate if not i.isdigit()]) \
                                     + str(int(''.join([i for i in row[i].coordinate if i.isdigit()])) + 1)
                        cell_value = data_wb[value_cell].value
                        if col_value < tracker_headers['Duration Start']:
                            freq_headers[cell_value] = [row[i].coordinate,
                                                        ''.join([i for i in row[i].coordinate if not i.isdigit()]),
                                                        row[i].value]
                        else:
                            dur_headers[cell_value] = [row[i].coordinate,
                                                       ''.join([i for i in row[i].coordinate if not i.isdigit()]),
                                                       row[i].value]
            if row[i].value in tracker_headers:
                cell_value = row[i].value
                if cell_value == 'Session Data':
                    tracker_headers['Session Data Start'] = ''.join([i for i in row[i].coordinate if not i.isdigit()])
                elif cell_value == 'Frequency':
                    tracker_headers['Frequency Start'] = ''.join([i for i in row[i].coordinate if not i.isdigit()])
                elif cell_value == 'Duration':
                    tracker_headers['Duration Start'] = ''.join([i for i in row[i].coordinate if not i.isdigit()])
                elif cell_value == 'Session Time':
                    tracker_headers['Session Time Start'] = ''.join([i for i in row[i].coordinate if not i.isdigit()])
                elif cell_value == 'Pause Time':
                    tracker_headers['Pause Time Start'] = ''.join([i for i in row[i].coordinate if not i.isdigit()])
                tracker_headers[cell_value][0] = row[i].coordinate
                tracker_headers[cell_value][1] = ''.join([i for i in row[i].coordinate if not i.isdigit()])
                tracker_headers[cell_value][2] = ''.join([i for i in row[i].coordinate if i.isdigit()])
                tracker_headers[cell_value][3] = 0
                if i + 1 < len(row):
                    # Check if next cell is MergedCell,
                    while type(row[i + 1]) is MergedCell:
                        tracker_headers[cell_value][3] += 1
                        i += 1
    return tracker_headers, freq_headers, dur_headers


def increment_cell(cell, i):
    # TODO: Can probably make this more generic by taking the mod of 26 and applying appropriately while iterating through cell
    # ord('Z') - ord('A') = 25, 26 is a rollover, take ord('Z') - ord(cell) to find remainder before rollover
    number = cell[-1]
    cell = cell[:-1]
    if len(cell) == 1:
        if ord(cell) + i > 90:
            diff = ((ord(cell) + i) - 90) - 1
            new_cell = 'A' + increment_char('A', diff)
        else:
            new_cell = increment_char(cell, i)
        return new_cell + number
    elif len(cell) == 2:
        if ord(cell[1]) + i > 90:
            diff = ((ord(cell[1]) + i) - 90) - 1
            new_cell = increment_char(cell[0], 1) + increment_char('A', diff)
        else:
            new_cell = cell[0] + increment_char(cell[1], i)
        return new_cell + number
    else:
        print("You're asking for too much")


def increment_char(character, i):
    return chr(ord(character) + i)


def populate_spreadsheet(patient_name, ksf_excel, prim_session_dir, output_dir):
    # Load the tracker spreadsheet
    wb = openpyxl.load_workbook(ksf_excel)
    ksf_file = ksf_excel[:-5] + ".json"
    # Isolate the Data sheet
    data_wb = wb['Data']
    # Load in all of the sessions
    sessions = get_sessions(prim_session_dir)
    # If output folder doesn't exist, then make it
    if not path.exists(output_dir):
        os.mkdir(output_dir)
    # Get the key cells in the spreadsheet
    tracker_headers, freq_headers, dur_headers = get_key_cells(data_wb)
    # Assign values in header
    data_wb[tracker_headers['Assessment:____________'][0]] = "Assessment: " + sessions[0]['Assessment Name']
    data_wb[tracker_headers['Client:________'][0]] = "Client: " + patient_name
    # We expect the
    row, col, sess = 5, tracker_headers['Session Data Start'], 1
    for session in sessions:
        key_freq, key_dur = get_keystroke_info(ksf_file, session)
        data_wb[tracker_headers['Session'][1] + str(row)].value = sess
        data_wb[tracker_headers['Cond.'][1] + str(row)].value = session['Condition Name']
        data_wb[tracker_headers['Date'][1] + str(row)].value = session['Session Date']
        data_wb[tracker_headers['Therapist'][1] + str(row)].value = session['Session Therapist']
        data_wb[tracker_headers['Primary'][1] + str(row)].value = session['Primary Therapist']
        # data_wb[tracker_headers['Session'] + str(row)].value = session['Primary Data']
        data_wb[tracker_headers['Sess. Dur. (mins)'][1] + str(row)].value = int(session['Session Time']) / 60
        data_wb[tracker_headers['ST'][1] + str(row)].value = session['Session Time']
        data_wb[tracker_headers['PT'][1] + str(row)].value = session['Pause Time']
        # Populate frequency and duration keys
        for freq in key_freq:
            data_wb[freq_headers[freq][1] + str(row)].value = key_freq[freq]
        for dur in key_dur:
            data_wb[dur_headers[dur][1] + str(row)].value = key_dur[dur]
        row += 1
        sess += 1
    output_file = path.join(output_dir, f"{pathlib.Path(ksf_file).stem}_Charted.xlsx")
    try:
        wb.save(output_file)
    except PermissionError as e:
        messagebox.showerror("Error", "Permission denied to save charted tracker, make sure your have permissions to "
                                      f"save in the output directory or that the spreadsheet is closed!\n{output_file}")
        print("ERROR: Permission denied to save charted tracker, make sure your have permissions to "
              f"save in the output directory or that the spreadsheet is closed!\n{output_file}")
    return output_file


def get_keystroke_window(freq_bindings, dur_bindings, session_file, window_size):
    session_time = int(session_file["Session Time"])
    event_history = session_file["Event History"]
    if session_time % window_size != 0:
        session_time += window_size - (session_time % window_size)
    freq_windows = [[0] * len(freq_bindings) for i in range(int(session_time / window_size))]
    dur_windows = [[0] * len(dur_bindings) for i in range(int(session_time / window_size))]
    dur_keys, freq_keys = [], []
    for event in event_history:
        if type(event[1]) is list:
            dur_keys.append(event)
        else:
            freq_keys.append(event)
    for i in range(0, session_time, window_size):
        for event in freq_keys[:]:
            if i <= event[1] < (i + window_size):
                freq_windows[int(i / window_size)][freq_bindings.index(event[0])] += 1
                freq_keys.remove(event)
            else:
                break
    for i in range(0, session_time, window_size):
        for event in dur_keys[:]:
            if i <= event[1][0] < (i + window_size):
                dur_windows[int(i / window_size)][dur_bindings.index(event[0])] += 1
                if i <= event[1][1] < (i + window_size):
                    dur_keys.remove(event)
            elif i <= event[1][1] < (i + window_size):
                dur_windows[int(i / window_size)][dur_bindings.index(event[0])] += 1
                dur_keys.remove(event)
            else:
                break
    return freq_windows, dur_windows


def get_keystroke_info(key_file, session_file):
    freq_bindings = {}
    dur_bindings = {}
    key_freq = {}
    key_dur = {}
    with open(key_file) as f:
        keystroke_json = json.load(f)
    for key in keystroke_json:
        if key == "Frequency":
            for bindings in keystroke_json[key]:
                freq_bindings[bindings[1]] = bindings[0]
                key_freq[bindings[1]] = 0
        if key == "Duration":
            for bindings in keystroke_json[key]:
                dur_bindings[bindings[1]] = bindings[0]
                key_dur[bindings[1]] = 0
    event_history = session_file["Event History"]
    for session_info in event_history:
        session_param = session_info[0]
        try:
            if session_param in freq_bindings:
                key_freq[session_param] += 1
            elif session_param in dur_bindings:
                key_dur[session_param] += int(session_info[1][1]) - int(session_info[1][0])
        except Exception as e:
            print(f"ERROR: Error encountered\n{str(e)}\n{traceback.print_exc()}")
            continue
    return key_freq, key_dur


def get_export_files(export_dir):
    export_files = []
    if path.isdir(export_dir):
        _, _, files = next(walk(export_dir))
        for file in files:
            if pathlib.Path(file).suffix == ".csv":
                export_files.append(pathlib.Path(file).stem)
    return export_files


def get_session_files(session_dir):
    session_files = []
    if path.isdir(session_dir):
        _, _, files = next(walk(session_dir))
        for file in files:
            if pathlib.Path(file).suffix == ".json":
                session_files.append(path.join(session_dir, file))
    return session_files


def get_sessions(session_dir):
    session_files = []
    sessions = []
    if path.isdir(session_dir):
        _, _, files = next(walk(session_dir))
        for file in files:
            if pathlib.Path(file).suffix == ".json":
                session_files.append(file)
    for file in session_files:
        with open(path.join(session_dir, file), 'r') as f:
            data = json.load(f)
            if data['Primary Data'] == "Primary":
                sessions.append(data)
    return sessions


def open_keystroke_file(key_file):
    bindings = []
    with open(key_file) as f:
        keystroke_json = json.load(f)
    for key in keystroke_json:
        if key != "Name":
            bindings.append((key, keystroke_json[key]))
    return bindings


def create_new_ksf_revision(original_ksf, keystrokes):
    ksf_wb = openpyxl.load_workbook(original_ksf)
    data_wb = ksf_wb['Data']
    cond_wb = ksf_wb['Conditions']
    tracker_headers, freq_headers, dur_headers = get_key_cells(data_wb)
    wb = Workbook()
    ws = wb.active
    wb.create_sheet("Conditions")
    cond_ws = wb['Conditions']
    ws.title = "Data"
    skip_headers = ['Session Data', 'Frequency', 'Duration', 'ST', 'PT', 'Session Time', 'Pause Time']
    for key in tracker_headers:
        if key in skip_headers:
            continue
        if type(tracker_headers[key]) is list:
            ws[tracker_headers[key][0]].value = key
            if tracker_headers[key][3] > 0:
                cell_range = f"{tracker_headers[key][0]}:" \
                             f"{increment_char(tracker_headers[key][1], tracker_headers[key][3])}{tracker_headers[key][2]}"
                ws.merge_cells(cell_range)

    for key in freq_headers:
        ws[freq_headers[key][0]].value = freq_headers[key][2]
        name_cell = freq_headers[key][1] + '4'
        ws[name_cell].value = key

    if len(freq_headers) != len(keystrokes['Frequency']):
        difference = len(keystrokes['Frequency']) - len(freq_headers)
        last_cell = freq_headers[key][0]
        i = 1
        for key in keystrokes['Frequency'][-difference:]:
            ws[increment_cell(last_cell, i)].value = key[0]
            name_cell = increment_cell(last_cell, i)[:-1] + '4'
            ws[name_cell].value = key[1]
            i += 1
        for key in dur_headers:
            dur_headers[key][0] = increment_cell(dur_headers[key][0], difference)
        tracker_headers['Session Data'][3] += difference
        tracker_headers['Frequency'][3] += difference
        tracker_headers['Duration'][0] = increment_cell(tracker_headers['Duration'][0], difference)
        tracker_headers['ST'][0] = increment_cell(tracker_headers['ST'][0], difference)
        tracker_headers['PT'][0] = increment_cell(tracker_headers['PT'][0], difference)
        tracker_headers['Session Time'][0] = increment_cell(tracker_headers['Session Time'][0], difference)
        tracker_headers['Pause Time'][0] = increment_cell(tracker_headers['Pause Time'][0], difference)

    for key in dur_headers:
        ws[dur_headers[key][0]].value = dur_headers[key][2]
        name_cell = dur_headers[key][0][:-1] + '4'
        ws[name_cell].value = key

    if len(dur_headers) != len(keystrokes['Duration']):
        difference = len(keystrokes['Duration']) - len(dur_headers)
        last_cell = dur_headers[key][0]
        i = 1
        for key in keystrokes['Duration'][-difference:]:
            ws[increment_cell(last_cell, i)].value = key[0]
            name_cell = increment_cell(last_cell, i)[:-1] + '4'
            ws[name_cell].value = key[1]
            i += 1
        tracker_headers['Session Data'][3] += difference
        tracker_headers['Duration'][3] += difference
        tracker_headers['ST'][0] = increment_cell(tracker_headers['ST'][0], difference)
        tracker_headers['PT'][0] = increment_cell(tracker_headers['PT'][0], difference)
        tracker_headers['Session Time'][0] = increment_cell(tracker_headers['Session Time'][0], difference)
        tracker_headers['Pause Time'][0] = increment_cell(tracker_headers['Pause Time'][0], difference)

    for key in skip_headers:
        if type(tracker_headers[key]) is list:
            ws[tracker_headers[key][0]].value = key
            if tracker_headers[key][3] > 0:
                cell_range = f"{tracker_headers[key][0]}:" \
                             f"{increment_char(tracker_headers[key][0][:-1], tracker_headers[key][3])}{tracker_headers[key][2]}"
                ws.merge_cells(cell_range)
    for i in range(1, 20):
        if cond_wb['A' + str(i)].value is None:
            break
        else:
            cond_ws['A' + str(i)].value = cond_wb['A' + str(i)].value
    ksf_dir = pathlib.Path(original_ksf).parent
    ksf_count = len(glob.glob1(ksf_dir, "*.xlsx"))
    if ksf_count > 1:
        new_ksf = f"{original_ksf[:-8]}_V{ksf_count + 1}.xlsx"
    else:
        new_ksf = f"{original_ksf[:-5]}_V{ksf_count + 1}.xlsx"
    wb.save(new_ksf)
    return new_ksf


def compare_keystrokes(old, new):
    if old['Frequency'] == new['Frequency']:
        if old['Duration'] == new['Duration']:
            if old['Conditions'] == new['Conditions']:
                return True
    return False
